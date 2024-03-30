import openpyxl
import tkinter.filedialog
import os
import datetime


def time():
    return datetime.datetime.now().strftime('%y%m%d_%H%M%S')


def ex_search_1():
    # 동일 디렉토리에 있는 xlsx 파일 중 첫 번째 파일 선택
    file = os.listdir(os.getcwd())
    return [s for s in file if '.xlsx' in s][0]


def ex_search_2():
    # 탐색기로 파일 선택
    return tkinter.filedialog.askopenfilename(initialdir='', title='한 개의 파일 선택')


def file_read():
    # a = input('1 : 자동 실행, 2 : 직접 선택\n')
    a = '2'
    # file read
    if a == '1':
        return openpyxl.load_workbook(ex_search_1())
    elif a == '2':
        return openpyxl.load_workbook(ex_search_2())
    else:
        return print('잘못된 선택')


# def file_read_2():
#     # 여러 sheet를 읽을 경우
#     for sheet_nm in file_read().sheetnames:
#         sheet = file_read()[sheet_nm]


def func_connect():
    file = file_read()
    file_sheet = file['TC']
    for row_data in file_sheet.iter_rows(min_row=2, min_col=0): # row 열, col 행
        i = 1  # 동적 변수 count를 위한 변수 선언
        for cell in row_data:
            globals()[f'cell_{i}'] = cell  # 동적 변수 생성
            #####################
            # if i == 6: # i와 비교하는 숫자를 능동적으로 대응할 수 있도록 하는 code로 업데이트 검토 필요 (변수 삽입이 완료된 이후에 입력을 시작하도록 의도함)
            if i == len(row_data):
                # cell_2.value : function name
                # cell_6.value : parameter
                # cell_7.value : Expected Result
                # cell_8.value : Test Result
                print(globals()[cell.value](cell_6.value))
                # print(cell_6.value)
                # print(cell_7.value)
                # print(cell_8.value)

                if globals()[cell_2.value](cell_6.value) == cell_7.value:
                    file_sheet[cell_8.coordinate] = 'PASS'
                else:
                    file_sheet[cell_8.coordinate] = 'FAIL'
            # 행을 순차적으로 증가시키기 위한 증가
            i += 1
            #####################

    file_name = f'TC_result_{datetime.time()}.xlsx'
    file.save(file_name)  # file Save 동작은 for문 밖으로 실행해야 함 (for 문 안에서 할 경우 여러 파일 생성됨)
    file.close()
    print('작성 완료\nFile Name : ', file_name)


if __name__ == '__main__':
    func_connect()
