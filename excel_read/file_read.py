import openpyxl
import tkinter.filedialog
import os, json

import selenium.webdriver.common.devtools.v129.dom
from cffi.model import global_lock
from selenium.webdriver.common.devtools.v85.runtime import global_lexical_scope_names


# file read
def file_read(a='1'):
    # a = input('1 : 자동 실행, 2 : 직접 선택\n')
    if a == '1':  # 동일 디렉토리에 있는 xlsx 파일 중 첫 번째 파일 선택
        file = os.listdir(os.getcwd())
        return openpyxl.load_workbook([s for s in file if '.xlsx' in s][0])
    elif a == '2':  # 탐색기로 파일 선택
        ################## TBD ##################
        return openpyxl.load_workbook(tkinter.filedialog.askopenfilename(initialdir='', title='한 개의 파일 선택'))
    else:
        return print('잘못된 선택')


def sheet_read(a='1'):
    if a == '1':  # 특정 sheet를 읽을 경우
        return 'TC'
    else:   # 여러 sheet를 읽을 경우
        ################## TBD ##################
        for sheet_nm in file_read().sheetnames:
            sheet = file_read()[sheet_nm]
        return None


class FUNC:
    def __init__(self):
        self.file = file_read()
        self.file_sheet = self.file[sheet_read()]
        self.case_item = {}
        self.case_result = {}

    def test_item_array(self):
        i = 1  # 열
        j = 1
        for col_data in self.file_sheet.iter_rows(min_col=1, min_row=2, max_col=self.file_sheet.max_column,
                                                  max_row=self.file_sheet.max_row):
            for item_v in col_data:
                if (i + (1*j)) % self.file_sheet.max_column == 0 or i % self.file_sheet.max_column == 0:
                    globals()[f'em_{i}'] = item_v.coordinate
                else:
                    globals()[f'em_{i}'] = '' if item_v.value is None else item_v.value

                if i % self.file_sheet.max_column == 0:  # 행의 모든 열을 변수에 삽입 후 진입
                    z = i - (self.file_sheet.max_column - 1)  # 두 번째 행(row) 부터 읽기 때문에 -1 추가
                    self.case_item[globals()[f"em_{z + 0}"]] = {
                        "function_name": globals()[f'em_{z + 1}'],
                        "number": globals()[f'em_{z + 2}'],
                        "idx": globals()[f'em_{z + 3}'],
                        "input_1": globals()[f'em_{z + 6}'],
                        "input_2": globals()[f'em_{z + 7}'],
                        "input_3": globals()[f'em_{z + 8}'],
                        "input_4": globals()[f'em_{z + 9}']
                    }
                    self.case_result[globals()[f"em_{z + 0}"]] = {
                        "function_name": globals()[f'em_{z + 1}'],
                        "number": globals()[f'em_{z + 2}'],
                        "Expected_result": globals()[f'em_{z + 10}'],
                        "Actual_result": globals()[f'em_{z + 11}'],  # 결과 기록 위치 저장
                        "record": globals()[f'em_{z + 12}']  # 결과 기록 위치 저장
                    }
                i += 1
        return json.dumps({
            "item": self.case_item,
            "result": self.case_result
        })

    def test_result_write(self, result_cell, record_cell, result, record=None):
        print('1', result_cell)
        print('2', record_cell)
        print('3', result)
        print('4', type(record))

if __name__ == "__main__":
    run = FUNC()
    print(type(run.test_item_array()))